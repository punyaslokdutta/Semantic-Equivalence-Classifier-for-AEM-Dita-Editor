# -*- coding: utf-8 -*-
"""
Created on Tue May 23 14:40:10 2017

@author: pudutta
"""

from __future__ import absolute_import, division, print_function
import codecs
import glob
import multiprocessing
import os
import pprint
import re
import nltk
import gensim.models.word2vec as w2v
import math
import sklearn.manifold
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
from itertools import chain 
from scipy import spatial
from collections import Counter
from operator import itemgetter
from xlwt import Workbook
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

threshold=94
get=0
na=0
workbook = load_workbook(filename='C:/Users/pudutta/Desktop/trainquora2.xlsx')
first_sheet ="Mysheet"
ws = workbook.get_sheet_by_name(first_sheet)
for row in range(2, 10000):
    if(ws.cell(row=row, column=4).value=="NA"):
        na=na+1
    else:    
        if((ws.cell(row=row, column=5).value>threshold and ws.cell(row=row, column=3).value==1) or (ws.cell(row=row, column=5).value<threshold and ws.cell(row=row, column=3).value==0) ):
            print((ws.cell(row=row, column=1).value))
            print((ws.cell(row=row, column=2).value))
            print(row)
            get=get+1
print(get/(9999-na))        
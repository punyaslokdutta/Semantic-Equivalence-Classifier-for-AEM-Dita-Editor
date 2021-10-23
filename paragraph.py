# -*- coding: utf-8 -*-
"""
Created on Mon May 22 19:12:58 2017

@author: pudutta
"""

from __future__ import absolute_import, division, print_function
import codecs
import glob
import multiprocessing
import os
import pprint
import re
import nltk
import gensim.models.word2vec as w2v
import math
import sklearn.manifold
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
from itertools import chain 
from scipy import spatial
from collections import Counter
from operator import itemgetter
from xlwt import Workbook
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

workbook = load_workbook(filename='C:/Users/pudutta/Desktop/sample.xlsx')
f=open("C:/Users/pudutta/Desktop/Paragraph.txt", "w", encoding="utf-8")
first_sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(first_sheet)
string1="A"
string2="B"
for  row in range(1, 501):
    sentence1=worksheet[string1+str(row)].value
    sentence1=sentence1.lower()                    
    sentence2=worksheet[string2+str(row)].value
    sentence2=sentence2.lower() 
    f.write(sentence1+"\n")
    f.write(sentence2+"\n")


                   
   
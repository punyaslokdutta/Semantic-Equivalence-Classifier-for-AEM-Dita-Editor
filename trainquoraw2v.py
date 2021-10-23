# -*- coding: utf-8 -*-
"""
Created on Thu Jun 29 14:52:12 2017

@author: pudutta
"""

# -*- coding: utf-8 -*-

from __future__ import absolute_import, division, print_function
import codecs
import glob
import multiprocessing
import os
import pprint
import re
import nltk
import gensim.models.word2vec as w2v
import math
import sklearn.manifold
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
from itertools import chain 
from scipy import spatial
from collections import Counter
from operator import itemgetter
from xlwt import Workbook
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


nltk.download("punkt")
nltk.download("stopwords")
lemmatizer = WordNetLemmatizer()

def sentence_to_wordlist(raw):
    clean = re.sub("[^a-zA-Z]"," ", raw)
    words = clean.split()
    words = [w for w in words if not w in stopwords.words("english")]
    for w in words:
        lemmatizer.lemmatize(w)
       
    return words 


words=[]

#gloveFile = 'C:/Users/pudutta/Desktop/gloves/glove2.txt'    
#def loadGloveModel(gloveFile):
    #print ("Loading Glove Model")
    #f = open(gloveFile,'r',encoding='utf-8')
   # model = {}
    #for line in f:
       # splitLine = line.split()
        #word = splitLine[0]
       # embedding = [float(val) for val in splitLine[1:]]
        #model[word] = embedding
   # print ("Done")
    #return model


#model=loadGloveModel(gloveFile)
#print(".........")
#print(model['2016'])
#print('2016')
lemmatizer = WordNetLemmatizer()
num_features = 300
min_word_count = 1
num_workers = multiprocessing.cpu_count()
context_size = 5
downsampling = 1e-3
seed=1

#thrones2vec=gensim.models.Word2Vec( size=300, window=5, min_count=1, workers=num_workers, seed=seed, sg=1,sample=downsampling)
thrones2vec = w2v.Word2Vec(
        sg=1,
        seed=seed,
        workers=num_workers,
        size=num_features,
        min_count=min_word_count,
        window=context_size,
        sample=downsampling
        )

def tokenize_sentence( sentence):
        clean = re.sub("[^a-zA-Z0-9]"," ", sentence)
        words = clean.split()
        #tokens = nltk.word_tokenize(sentence)
        words = [w for w in words if not w in stopwords.words("english")]
        #print(tokens)
        return words  


workbook = load_workbook(filename='C:/Users/pudutta/Desktop/trainquora2.xlsx')
wb = Workbook()
ws = wb.create_sheet("Mysheet", 0) 
first_sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(first_sheet)
string1='A'
string2='B'
string3='C'
string4='D'
string5='E'
words1=[]
words2=[]
na=0
error=0
noninf=0
get=0
sentences=[]
for row in range(2, 10000):
    sentence1=worksheet[string1+str(row)].value
    sentence1=sentence1.lower() 
    tokens1=tokenize_sentence(sentence1)
    sentences.append(tokens1)                   
    sentence2=worksheet[string2+str(row)].value
    sentence2=sentence2.lower()
    tokens2=tokenize_sentence(sentence2)
    sentences.append(tokens2)
    
thrones2vec.build_vocab(sentences)
thrones2vec.train(sentences)  
cosinevaluesx=[0.0]*10000  
maxi=0
mini=101
for  row in range(2, 10000):
        errorhere=0
        ws.cell(row=row, column=1).value=worksheet.cell(row=row, column=1).value
        ws.cell(row=row, column=2).value=worksheet.cell(row=row, column=2).value 
        ws.cell(row=row, column=3).value=worksheet.cell(row=row, column=3).value  
        sentence1=worksheet[string1+str(row)].value
        sentence1=sentence1.lower() 
        #sentences.append(sentence1)                   
        sentence2=worksheet[string2+str(row)].value
        sentence2=sentence2.lower()
        #sentences.append(sentence2)                    
        words1=sentence_to_wordlist(sentence1)  
        words2=sentence_to_wordlist(sentence2)
        #worksheet.cell(row=row, column=6).value=25*(worksheet2.cell(row=row, column=1).value)
       #workbook.save("C:/Users/pudutta/Desktop/trainquora.xlsx")
        #print(words1)
        #print(words2)
        xx=1
        xy=1
        
           
        for idx, words in enumerate(words1):
            if(words in thrones2vec.wv.vocab):
                continue
            else:
                xx=0
                    #print(words)
        for idx, words in enumerate(words2):
            if(words in thrones2vec.wv.vocab):
                continue
            else:
                xy=0
                    #print(words)
            
        if(xx==0 or xy==0 or len(words1)==0 or len(words2)==0):
                #print(row)
                ws.cell(row=row, column=4).value="NA"
                ws.cell(row=row, column=5).value="NA"
                na=na+1
                cosinevaluesx[row]=-1
        else:
                sum1=[0]*300
                for idx, words in enumerate(words1):
                    sum1=[sum(x) for x in zip(sum1, thrones2vec.wv[words] )]
              
                sum2=[0]*300    
                for idx, words in enumerate(words2):
                    sum2=[sum(x) for x in zip(sum2, thrones2vec.wv[words])]
                for idx, x in enumerate(sum1):
                    sum1[idx]=sum1[idx]/len(words1)
            
                for idx, x in enumerate(sum2):
                    sum2[idx]=sum2[idx]/len(words2)
                cosvalue=100*(1 - spatial.distance.cosine(sum1, sum2))
                ws.cell(row=row, column=5).value=cosvalue
                maxi=max(maxi,cosvalue )
                mini=min(mini, cosvalue)
                cosinevaluesx[row]=cosvalue
                
                
for row in range(2, 10000):
    if(ws.cell(row=row, column=5).value!="NA"):
        ws.cell(row=row, column=5).value=(ws.cell(row=row, column=5).value-mini)*(100/(maxi-mini))

#print(get/(10000-na))       
wb.save("C:/Users/pudutta/Desktop/trainquoraw2vnormalised.xlsx") 
workbook.save("C:/Users/pudutta/Desktop/trainquora2.xlsx")             
         
         
         
         
         
         